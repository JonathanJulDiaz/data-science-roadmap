import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename : str) -> None:

    """
    Process a workbook by applying a 10% discount to all products in
    the 'Price' column, and then creating a bar chart showing the
    corrected prices.

    The workbook is then saved back to the same file. The chart is
    placed in the 'E2' cell of the first sheet. The corrected prices
    are placed in a new column next to the original price column.

    :param filename: The name of the Excel file to process.
    :return: None
    """
    # Load the workbook and select the first sheet
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # Iterate over each row to apply a 10% discount to the price in column 3
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row=row, column=4)
        corrected_price_cell.value = corrected_price

    # Add a header for the corrected price column
    corrected_price_cell = sheet.cell(row=1, column=4)
    corrected_price_cell.value = 'Corrected Price'

    # Create a reference for the corrected prices to include in the chart
    values = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=sheet.max_row)

    # Create a bar chart
    chart = BarChart()
    chart.add_data(values, titles_from_data=True)
    chart.title = 'Corrected Price'
    chart.x_axis.title = 'Product'
    chart.y_axis.title = 'Price'

    # Add the chart to the sheet at position E2
    sheet.add_chart(chart, 'E2')

    # Save the modified workbook back to the same file
    wb.save(filename)
